import openpyxl
import sqlite3
import os
import re

# Resolve paths relative to this script
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "5.6.2569  อาจารย์  ส่งมาให้ล่าสุด WIC TSK Team.xlsx")
db_path = os.environ.get("DB_PATH", os.path.join(script_dir, "wisco_orders.db"))

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace(u'\xa0', u' ')
    return s

def clean_float(val):
    if val is None:
        return 0.0
    s = str(val).strip()
    s = re.sub(r'[^\d\.\-]', '', s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0

def main():
    print("Connecting to unified SQLite database...")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Drop and recreate table for clean seeding
    cursor.execute("DROP TABLE IF EXISTS customers;")
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS customers (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      customer_name TEXT NOT NULL,
      wic_customer_id TEXT,
      tier TEXT,
      status TEXT,
      registration_date TEXT,
      sale_code TEXT,
      team TEXT,
      month_forecast REAL DEFAULT 0,
      year_forecast REAL DEFAULT 0,
      zone TEXT,
      industry TEXT,
      product_service TEXT,
      contact_name TEXT,
      position TEXT,
      tel TEXT,
      mobile TEXT,
      line_id TEXT,
      email TEXT,
      address TEXT,
      subdistrict TEXT,
      district TEXT,
      province TEXT,
      zipcode TEXT,
      last_visit_date TEXT DEFAULT '',
      last_call_date TEXT DEFAULT '',
      last_contact_date TEXT DEFAULT '',
      next_planned_date TEXT DEFAULT '',
      next_planned_type TEXT DEFAULT '',
      created_at TEXT DEFAULT (datetime('now','localtime')),
      updated_at TEXT DEFAULT (datetime('now','localtime'))
    );
    """)
    
    # Ensure activity/plan tables exist
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS activities (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      customer_id INTEGER NOT NULL,
      activity_type TEXT NOT NULL,
      activity_date TEXT NOT NULL,
      contact_person TEXT DEFAULT '',
      summary TEXT DEFAULT '',
      created_by TEXT DEFAULT '',
      created_at TEXT DEFAULT (datetime('now','localtime')),
      FOREIGN KEY(customer_id) REFERENCES customers(id) ON DELETE CASCADE
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS plans (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      customer_id INTEGER NOT NULL,
      plan_type TEXT NOT NULL,
      planned_date TEXT NOT NULL,
      objective TEXT DEFAULT '',
      created_by TEXT DEFAULT '',
      created_at TEXT DEFAULT (datetime('now','localtime')),
      FOREIGN KEY(customer_id) REFERENCES customers(id) ON DELETE CASCADE
    );
    """)
    
    print("Opening Excel file...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = 'รวม TSK '
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found!")
        return
        
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    data_rows = rows[2:]
    
    inserted_count = 0
    for idx, row in enumerate(data_rows):
        cust_name = clean_str(row[1]) if len(row) > 1 else ""
        if not cust_name:
            continue
            
        wic_id = clean_str(row[2]) if len(row) > 2 else ""
        tier = clean_str(row[3]) if len(row) > 3 else ""
        status = clean_str(row[4]) if len(row) > 4 else ""
        reg_date = clean_str(row[5]) if len(row) > 5 else ""
        sale_code = clean_str(row[6]) if len(row) > 6 else ""
        team = clean_str(row[7]) if len(row) > 7 else ""
        month_fc = clean_float(row[8]) if len(row) > 8 else 0.0
        year_fc = clean_float(row[9]) if len(row) > 9 else 0.0
        zone = clean_str(row[10]) if len(row) > 10 else ""
        industry = clean_str(row[11]) if len(row) > 11 else ""
        prod_svc = clean_str(row[12]) if len(row) > 12 else ""
        contact = clean_str(row[13]) if len(row) > 13 else ""
        pos = clean_str(row[14]) if len(row) > 14 else ""
        tel = clean_str(row[15]) if len(row) > 15 else ""
        mobile = clean_str(row[16]) if len(row) > 16 else ""
        
        lid1 = clean_str(row[17]) if len(row) > 17 else ""
        lid2 = clean_str(row[18]) if len(row) > 18 else ""
        line_id = lid1 if lid1 else lid2
        
        email = clean_str(row[19]) if len(row) > 19 else ""
        addr = clean_str(row[21]) if len(row) > 21 else ""
        subdist = clean_str(row[22]) if len(row) > 22 else ""
        dist = clean_str(row[23]) if len(row) > 23 else ""
        prov = clean_str(row[24]) if len(row) > 24 else ""
        zipc = clean_str(row[25]) if len(row) > 25 else ""
        
        if prov:
            prov = prov.replace("จ.", "").strip()
        if not prov and zone:
            z_clean = zone.replace("จ.", "").strip()
            if z_clean in ["กรุงเทพ", "กทม.", "สมุทรปราการ", "นนทบุรี", "นครปฐม", "สมุทรสาคร", "ระยอง", "ฉะเชิงเทรา", "ชลบุรี", "ปราจีนบุรี", "ปทุมธานี", "สระบุรี"]:
                prov = z_clean
        
        if prov == "กทม.":
            prov = "กรุงเทพ"
            
        cursor.execute("""
        INSERT INTO customers (
            customer_name, wic_customer_id, tier, status, registration_date,
            sale_code, team, month_forecast, year_forecast, zone,
            industry, product_service, contact_name, position, tel,
            mobile, line_id, email, address, subdistrict,
            district, province, zipcode
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (
            cust_name, wic_id, tier, status, reg_date,
            sale_code, team, month_fc, year_fc, zone,
            industry, prod_svc, contact, pos, tel,
            mobile, line_id, email, addr, subdist,
            dist, prov, zipc
        ))
        inserted_count += 1
        
    conn.commit()
    conn.close()
    print(f"Successfully seeded database with {inserted_count} customers!")

if __name__ == "__main__":
    main()
